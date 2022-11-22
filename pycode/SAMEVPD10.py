# -*- coding: utf-8 -*-
"""
Created on Mon Nov 21 22:59:42 2022

@author: Lenovo
"""


'''
统计同一VPD下的Ttemperature和RESP数据,通过不同的vpd计算得到不同的温度敏感性，看温度敏感性的变化
此版本换份温度围，划分地表植被类型，对所有的同种植被类型整合，进行回归
'''


import pandas as pd
import os
import numpy as np
from sklearn import linear_model 
import openpyxl


#----------------------------------------------------------------------------------------------------------------------------------
def TREdata(fileN):      #d读取源文件中的VPD、T和RESP
    eco = pd.read_csv(fileN);
    ecom = len(eco);
    data = np.zeros((1,3),dtype = float);
    a = np.zeros((1,3));
    n1 = 0;
    for i1 in range(ecom):
       if eco['TA_F_MDS'][i1]!=-9999 and eco['VPD_F_MDS'][i1]!=-9999 and eco['RECO_NT_VUT_REF'][i1]!=-9999 and eco['RECO_NT_VUT_REF'][i1]!=0:
           data[n1][0] = 1000/(eco['TA_F_MDS'][i1]+273);
           data[n1][1] = eco['VPD_F_MDS'][i1];
           data[n1][2] = np.log(eco['RECO_NT_VUT_REF'][i1]);
           data = np.row_stack((data,a));
           n1 = n1 + 1;
    return data       #第一列T，第二列VPD，第三列RESP
#----------------------------------------------------------------------------------------------------------------------------------



#----------------------------------------------------------------------------------------------------------------------------------
def SameTvalue(data,T0,Tn,dT):          #读取同一T范围内的T和RESP  T0 = -5, Tn = 35, dT = 10
    l = int((Tn - T0)/dT + 1);
    ecom = len(data);
    value = np.zeros((10,3*l),dtype = float);
    a = np.zeros((1,3*l));

    T = T0;
    for i0 in range(l):
        value[0][i0*3] = T;
        T = T + dT;
    
    for i1 in range(ecom):
        for i2 in range(l):
            if data[i1][0] >= value[0][i2*3] - dvpd/2 and data[i1][0] < value[0][i2*3] + dvpd/2:
                value[1][i2*3] = value[1][i2*3] + 1;
                b = list(value[1,...]);
                mm = max(b);
                [m,n] = value.shape;
                if mm+2 > m:
                    value = np.row_stack((value,a));
                value[int(value[1][i2*3])+ 1][i2*3] = data[i1][1];
                value[int(value[1][i2*3])+ 1][i2*3 + 1] = data[i1][0];
                value[int(value[1][i2*3])+ 1][i2*3 + 2] = data[i1][2];  
    return value        #第一列VPD，第二列T，第三列RESP
#----------------------------------------------------------------------------------------------------------------------------------



#----------------------------------------------------------------------------------------------------------------------------------
def part(data,i):                 #此处data来自于SameTvalue，将SameTvalue得到的数据拆分
    m =  data[1][i*3];            #i表示第i个温度范围
    tdata = np.zeros((int(m),3));
    for j in range(int(m)):
        tdata[j][0] = data[j+2][3*i];
        tdata[j][1] = data[j+2][3*i+1];
        tdata[j][2] = data[j+2][3*i+2];
    return tdata       #第一列VPD，第二列T，第三列RESP
#----------------------------------------------------------------------------------------------------------------------------------



#----------------------------------------------------------------------------------------------------------------------------------
def SameVPDvalue(data,vpd0,vpdn,dvpd):          #读取同一VPD范围内的T和RESP,data来自与part里的tdata
    l = int((vpdn - vpd0)/dvpd + 1);
    ecom = len(data);
    value = np.zeros((10,2*l),dtype = float);
    a = np.zeros((1,2*l));

    vpd = vpd0;
    for i0 in range(l):
        value[0][i0*2] = vpd;
        vpd = vpd + dvpd;
    
    for i1 in range(ecom):
        for i2 in range(l):
            if data[i1][0] >= value[0][i2*2] - dvpd/2 and data[i1][0] < value[0][i2*2] + dvpd/2:
                if data[i1][1] != -9999 and data[i1][2] !=0 and data[i1][2] != -9999:
                    value[1][i2*2] = value[1][i2*2] + 1;
                    b = list(value[1,...]);
                    mm = max(b);
                    [m,n] = value.shape;
                    if mm+2 > m:
                        value = np.row_stack((value,a));
                    value[int(value[1][i2*2])+ 1][i2*2] = data[i1][1];
                    value[int(value[1][i2*2])+ 1][i2*2 + 1] = data[i1][2];
    return value               #第一列T，第二列RESP
#----------------------------------------------------------------------------------------------------------------------------------



#----------------------------------------------------------------------------------------------------------------------------------
def regression(value):
    [m,n] = value.shape;
    slope = np.zeros((1,int(n/2)));
    for i in range(int(n/2)):
        a = [];
        b = [];
        for i1 in range(m-3):
            if value[i1+2][2*i]!=0 and value[i1+2][2*i+1]!=0:
                a.append(value[i1+2][2*i]);
                b.append(value[i1+2][2*i+1]);
        regr = linear_model.LinearRegression();
        a = np.array(a);
        b = np.array(b);
        if len(a) >= 100:
            regr.fit(a.reshape(-1,1),b);
            k = regr.coef_;
            slope[0][i] = -k[0]*8.62/100;
    return slope
#----------------------------------------------------------------------------------------------------------------------------------


#----------------------------------------------------------------------------------------------------------------------------------
#读取所需处理站点文件，返回需要读取的站点
def siteread(pathsite):   # path为想要读取的excel文档
    file = openpyxl.load_workbook(pathsite); #打开想要读取的excel文档
    sheet1 = file['Sheet1'];
    m = sheet1.max_row;
    n = sheet1.max_column;
    pathr = [];
    for i in range(n):
        if sheet1.cell(1,i+1).value == 'fluxnetid':
            for j in range(m-1):
                pathr.append(sheet1.cell(j+2,i+1).value);
        elif sheet1.cell(1,i+1).value == 'igbp_land_use':
            a = sheet1.cell(2,i+1).value;
        else:
            continue;
    return pathr, a         #a为植被类型，pathr为站点信息
#----------------------------------------------------------------------------------------------------------------------------------



#----------------------------------------------------------------------------------------------------------------------------------
pathinfo = 'D:/Data/fluxnet/TreatedData/ClassLandcover/LandCoverinfo/';  #储存分类信息excel的文件夹
path = 'D:/Data/fluxnet/OriginalData/AllHourlyData/';  #储存原始数据的位置

T0 = -5; 
Tn = 35;
dT = 10;
vpd0 = 5;
vpdn = 35;
dvpd = 10;
n1 = int((Tn - T0)/dT +1);
n2 = int((vpdn - vpd0)/dvpd + 1); 

E = addE = np.zeros((6,n2+2));
nE = 1;

for landinfo in os.listdir(pathinfo):
    pathsite = 'D:/Data/fluxnet/TreatedData/ClassLandcover/LandCoverinfo/' + landinfo;
    pathsiteid,a = siteread(pathsite);
    m = len(pathsiteid);
    
    E = np.zeros((2,n1*n2));           #储存活化能计算结果
    result = np.zeros((6, n1*n2*2));   #第一行温度范围，第二行vpd范围
    addre = np.zeros((1, n1*n2*2));

    for i in range(n1):
        E[0][n2*i] = T0 + i*dT;
        for j in range(n2):
            E[1][i*n2+j] = vpd0 + j*dvpd; 
    
    for i in range(n1):
        result[0][n2*2*i] = T0 + i*dT;
        for j in range(n2):
            result[1][i*n2*2+j*2] = vpd0 + j*dvpd; 

    for csv_file in os.listdir(path):
        for i in range(m):
            if csv_file[4:10] == pathsiteid[i]:
                dTRE = TREdata(path + csv_file);        ##读取原始数据
                dSameT = SameTvalue(dTRE,T0,Tn,dT);     ##统计同一温度范围下的T、RE和vpd                
                for i1 in range(n1):
                    parti1 =  part(dSameT,i1); 
                    dSameVPD = SameVPDvalue(parti1, vpd0, vpdn, dvpd);
                    [nds, mds] = dSameVPD.shape;
                    for i2 in range(n2):
                        for i3 in range(mds-2):
                            if dSameVPD[i3+2][i2*2] != 0 and dSameVPD[i3+2][i2*2+1] != 0:
                                result[2][i1*n2*2 + 2*i2] = result[2][i1*n2*2 + 2*i2] + 1;    #第三行用来储存该VPD范围有多少数据
                                b = list(result[2,...]);
                                mm = max(b);
                                [mre,nre] = result.shape;
                                if mm + 3 > mre:
                                    result = np.row_stack((result,addre));
                                result[ int(result[2][i1*n2*2 + 2*i2]) + 2 ][i1*n2*2 + 2*i2] = dSameVPD[i3+2][i2*2];
                                result[ int(result[2][i1*n2*2 + 2*i2]) + 2 ][i1*n2*2 + 2*i2 +1] = dSameVPD[i3+2][i2*2+1];

    Ei = regression(result);
    E = np.row_stack((E,Ei));

    #将统计的数据保存
    path00 = 'C:/Users/Lenovo/Desktop/未划分年份的初级数据/不同植被类型不同温度范围同一VPD的T及RESP/' + a + '.csv';
    result_pd = pd.DataFrame(result);
    result_pd.to_csv(path00, index= False, header= False)


path11 = 'C:/Users/Lenovo/Desktop/E.xlsx';
E_pd = pd.DataFrame(E);
writer2 = pd.ExcelWriter(path11);
E_pd.to_excel(writer2, sheet_name='sheet1');
#----------------------------------------------------------------------------------------------------------------------------------







