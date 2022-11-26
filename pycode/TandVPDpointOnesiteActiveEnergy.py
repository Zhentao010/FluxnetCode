
'''
统计同一VPD下的Ttemperature和RESP数据,通过不同的vpd计算得到不同的温度敏感性，看温度敏感性的变化
此版本不划分地表类型进行整合
'''

import pandas as pd
import os
import numpy as np
from sklearn import linear_model 
import openpyxl

#----------------------------------------------------------------------------------------------------------------------------------
''' 读取初始文件中的温度、vpd和呼吸速率数据，将无效数据进行剔除'''
def TREdata(fileN):      #d读取源文件中的VPD、T和RESP
    eco = pd.read_csv(fileN);
    ecom = len(eco);
    data = np.zeros((1,3),dtype = float);
    a = np.zeros((1,3));
    n1 = 0;
    for i1 in range(ecom):
        if eco['TA_F_MDS'][i1]!=-9999 and eco['VPD_F_MDS'][i1]!=-9999 and eco['RECO_NT_VUT_REF'][i1]!=-9999 and eco['RECO_NT_VUT_REF'][i1]!=0:
            data[n1][0] = eco['TA_F_MDS'][i1];
            data[n1][1] = eco['VPD_F_MDS'][i1];
            data[n1][2] = np.log(eco['RECO_NT_VUT_REF'][i1]);
            data = np.row_stack((data,a));
            n1 = n1 + 1;
    return data       #第一列T，第二列VPD，第三列RESP
#----------------------------------------------------------------------------------------------------------------------------------

#----------------------------------------------------------------------------------------------------------------------------------
''' 读取某一温度范围的T、vpd和呼吸速率数据'''
def SameTvalue(data,T0,Tn,dT,dt):          #读取同一T范围内的T和RESP
    l = int((Tn - T0)/dT + 1);             #dT表示所取的温度间隔，dt表示精度
    ecom = len(data);
    value = np.zeros((10,3*l),dtype = float);
    a = np.zeros((1,3*l));

    T = T0;
    for i0 in range(l):
        value[0][i0*3] = T;
        T = T + dT;
    
    for i1 in range(ecom):
        for i2 in range(l):
            if data[i1][0] >= value[0][i2*3] - dt/2 and data[i1][0] <= value[0][i2*3] + dt/2:
                value[1][i2*3] = value[1][i2*3] + 1;
                b = list(value[1,...]);
                mm = max(b);
                [m,n] = value.shape;
                if mm+2 > m:
                    value = np.row_stack((value,a));
                value[int(value[1][i2*3])+ 1][i2*3] = data[i1][1];
                value[int(value[1][i2*3])+ 1][i2*3 + 1] = data[i1][0];
                value[int(value[1][i2*3])+ 1][i2*3 + 2] = data[i1][2];  
    return value        #第一列VPD，第二列T，第三列RESP
#----------------------------------------------------------------------------------------------------------------------------------

#----------------------------------------------------------------------------------------------------------------------------------
'''将某一温度范围的数据提取出来供samevpddata函数处理 '''
def part(data,i):                 #此处data来自于SameTvalue，将SameTvalue得到的数据拆分
    m =  data[1][i*3];            #i表示第i个温度范围
    tdata = np.zeros((int(m),3));
    for j in range(int(m)):
        tdata[j][0] = data[j+2][3*i];
        tdata[j][1] = data[j+2][3*i+1];
        tdata[j][2] = data[j+2][3*i+2];
    return tdata       #第一列VPD，第二列T，第三列RESP
#----------------------------------------------------------------------------------------------------------------------------------

#----------------------------------------------------------------------------------------------------------------------------------
'''将提取到的某一温度范围的数据根据vpd范围进行划分 '''
def SameVPDvalue(data,vpd0,vpdn,dVPD,dvpd):          #读取同一VPD范围内的T和RESP,data来自与part里的tdata
    l = int((vpdn - vpd0)/dVPD + 1);                 ##dVPD为两温度间隔，dvpd为精度
    ecom = len(data);
    value = np.zeros((10,2*l),dtype = float);
    a = np.zeros((1,2*l));

    vpd = vpd0;
    for i0 in range(l):
        value[0][i0*2] = vpd;
        vpd = vpd + dVPD;
    
    for i1 in range(ecom):
        for i2 in range(l):
            if data[i1][0] >= value[0][i2*2] - dvpd/2 and data[i1][0] <= value[0][i2*2] + dvpd/2:
                if data[i1][1] != -9999 and data[i1][2] !=0 and data[i1][2] != -9999:
                    value[1][i2*2] = value[1][i2*2] + 1;
                    b = list(value[1,...]);
                    mm = max(b);
                    [m,n] = value.shape;
                    if mm+2 > m:
                        value = np.row_stack((value,a));
                    value[int(value[1][i2*2])+ 1][i2*2] = data[i1][1];
                    value[int(value[1][i2*2])+ 1][i2*2 + 1] = data[i1][2];
    return value               #第一列T，第二列RESP
#----------------------------------------------------------------------------------------------------------------------------------

#----------------------------------------------------------------------------------------------------------------------------------
def regression(a,b):
    regr = linear_model.LinearRegression();
    a = np.array(a);
    b = np.array(b);
    if len(a) >=10:
        regr.fit(a.reshape(-1,1),b);
        k = regr.coef_[0];
        Er = -k*8.62/100;
        B = regr.intercept_;
    else:
        Er = -9999;
        B = -9999;
    return Er, B
#----------------------------------------------------------------------------------------------------------------------------------

#----------------------------------------------------------------------------------------------------------------------------------
'''读取所需处理站点文件，返回需要读取的站点'''
def siteread(pathsite):   # path为想要读取的excel文档
    file = openpyxl.load_workbook(pathsite); #打开想要读取的excel文档
    sheet1 = file['Sheet1'];
    m = sheet1.max_row;
    n = sheet1.max_column;
    pathr = [];
    for i in range(n):
        if sheet1.cell(1,i+1).value == 'fluxnetid':
            for j in range(m-1):
                pathr.append(sheet1.cell(j+2,i+1).value);
        elif sheet1.cell(1,i+1).value == 'igbp_land_use':
            a = sheet1.cell(2,i+1).value;
        else:
            continue;
    return pathr, a         #a为植被类型，pathr为站点信息
#----------------------------------------------------------------------------------------------------------------------------------

#----------------------------------------------------------------------------------------------------------------------------------

'''求取平均值，最终结果放入average中'''

def average(result,T0,Tn,dT,vpd0,vpdn,dVPD):
    n1 = int((Tn - T0)/dT +1);
    n2 = int((vpdn - vpd0)/dVPD + 1); 
    average = np.zeros((n2+1, n1+1));
    
    for i in range(n1):
        average[0][i+1] = T0 + dT*i;
        
    for i in range(n2):
        average[i+1][0] = vpd0 + dVPD*i;
    
    [m,n] = result.shape;
    for i in range(n1):
        for j in range(n2):
            a = 0;
            if int(result[2][i*n2*2 +2*j]) >= 10:
                for k in range(int(result[2][i*n2*2 +2*j])):
                    a = a + result[k+3][i*n2*2 + 2*j+1];
                if int(result[2][i*n2*2 + 2*j]) != 0:
                    avera = a/int(result[2][i*n2*2 + 2*j]);
                    average[j+1][i+1] = avera;
    return average

#----------------------------------------------------------------------------------------------------------------------------------

'''
主函数
'''

#pathinfo = 'D:/Data/fluxnet/TreatedData/ClassLandcover/LandCoverinfo/';  #储存分类信息excel的文件夹
path = '/home/liu-z/main/fluxnet/originalData/AllHourlyData/';  #储存原始数据的位置

T0 = 30; 
Tn = 40;
dT = 1;
dt = 0.4;
vpd0 = 20;
vpdn = 41;
dVPD = 2;
dvpd = 2;
n1 = int((Tn - T0)/dT +1);
n2 = int((vpdn - vpd0)/dVPD + 1); 

E = np.zeros((1,n2));
addE = np.zeros((1,n2));
B = np.zeros((1,n2));
En = 0;

for csv_file in os.listdir(path):
    print(csv_file);
    result = np.zeros((6, n1*n2*2));   #第一行温度范围，第二行vpd范围
    addre = np.zeros((1, n1*n2*2));
    for i in range(n1):
        result[0][n2*2*i] = T0 + i*dT;
        for j in range(n2):
            result[1][n2*2*i+2*j] = vpd0 + j*dVPD; 
            
    dTRE = TREdata(path + csv_file);        ##读取原始数据
    dSameT = SameTvalue(dTRE,T0,Tn,dT,dt);     ##统计同一温度范围下的T、RE和vpd                
    for i1 in range(n1):
        parti1 =  part(dSameT,i1); 
        dSameVPD = SameVPDvalue(parti1, vpd0, vpdn, dVPD, dvpd);       ##统计同一VPD下的T和RESP
        [mds, nds] = dSameVPD.shape;
        for i2 in range(n2):
            for i3 in range(mds-2):
                if dSameVPD[i3+2][i2*2] != 0 and dSameVPD[i3+2][i2*2+1] != 0 and dSameVPD[i3+2][i2*2+1]>-10:
                    result[2][i1*n2*2 + 2*i2] = result[2][i1*n2*2 + 2*i2] + 1;    #第三行用来储存该VPD范围有多少数据
                    b = list(result[2,...]);
                    mm = max(b);
                    [mre,nre] = result.shape;
                    if mm + 3 > mre:
                        result = np.row_stack((result,addre));
                    result[ int(result[2][i1*n2*2 + 2*i2]) + 2 ][i1*n2*2 + 2*i2] = 1000/(dSameVPD[i3+2][i2*2] + 273 );
                    result[ int(result[2][i1*n2*2 + 2*i2]) + 2 ][i1*n2*2 + 2*i2 +1] = dSameVPD[i3+2][i2*2+1];

    #将统计的数据保存
    path00 = '/home/liu-z/main/fluxnet/TreatedData/单个站点VPD、T、RESP/30_40/原始数据/' + csv_file[4:10] + '.csv';
    result_pd = pd.DataFrame(result);
    result_pd.to_csv(path00, index= False, header= False);
    #将统计好的数据计算平均值
    path11 = '/home/liu-z/main/fluxnet/TreatedData/单个站点VPD、T、RESP/30_40/平均值/' + csv_file[4:10] + '.csv';
    daverage = average(result,T0,Tn,dT,vpd0,vpdn,dVPD);
    daverage_pd = pd.DataFrame(daverage);
    daverage_pd.to_csv(path11, index= False, header= False);
    for i in range(n1):
        daverage[0][i+1] = 1000/(daverage[0][i+1] + 273);
    for i in range(n2):
        Ea = [];
        Eb = [];
        for j in range(n1):
            if daverage[i+1][j+1] != 0:
                Ea.append(daverage[0][j+1]);
                Eb.append(daverage[i+1][j+1]);
        Ei,Bi = regression(Ea, Eb);
        E[En][i] = Ei;
        B[En][i] = Bi;
    En = En + 1;
    E = np.row_stack((E,addE));
    B = np.row_stack((B,addE));
    path22 = '/home/liu-z/main/fluxnet/TreatedData/单个站点VPD、T、RESP/30_40/E.csv';
    E_pd = pd.DataFrame(E);
    E_pd.to_csv(path22, index= False, header= False);
    path33 = '/home/liu-z/main/fluxnet/TreatedData/单个站点VPD、T、RESP/30_40/B.csv';
    B_pd = pd.DataFrame(B);
    B_pd.to_csv(path33, index= False, header= False);
    
#----------------------------------------------------------------------------------------------------------------------------------
