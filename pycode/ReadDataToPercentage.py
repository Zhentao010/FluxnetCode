import pandas as pd
import os
import numpy as np
import math
import xlwt
import openpyxl

def dayValue(fileN,variable):  #variable 为想要统计的量
    eco = pd.read_csv(fileN);
    ecom = len(eco);
    days = math.ceil(ecom/48)+1;
    dayValue = np.zeros((49,days),dtype = float);
    
    t = 0;
    for i0 in range(48):
        dayValue[i0+1][0] = t;
        if (i0+1)%2 == 1:
            t = t + 30;
        else:
            t = t +70;
    
    n = 1;
    for i1 in range(ecom):
        hour = eco['TIMESTAMP_START'][i1]%10000;
        day = math.floor(eco['TIMESTAMP_START'][i1]/10000);
        dayValue[0][n] = day;
        for i2 in range(48):
            if hour == dayValue[i2+1][0]:
                dayValue[i2+1][n] = eco[variable][i1];
            else:
                continue;
        if hour == 2330:
            n = n + 1;
        else:
            continue;
    return dayValue

def avermonth(dayValue):
    daym = len(dayValue);
    dayn = int(dayValue.size/daym);
    adday = np.zeros((daym,1));
    dayValue = np.insert(dayValue, dayn, adday, axis = 1);
    avermon = np.zeros((49,math.ceil((dayn-1)/365)*12+1));
    
    i0 = 1;
    t = 0;
    for i in range(48):
        avermon[i+1][0] = t;
        if (i+1)%2 == 1:
            t = t +30;
        else:
            t = t +70;
    i1 = 0;
    for i in range(dayn-1):
        for j in range(12):
            if math.floor((dayValue[0][i+1]%10000)/100) == j+1:
                i1 = i1+1;
                avermon[0][(i0-1)*12+j+1] = math.floor(dayValue[0][i+1]/100);
                for k in range(daym-1):
                    avermon[k+1][(i0-1)*12+j+1] = avermon[k+1][(i0-1)*12+j+1] + dayValue[k+1][i+1];
            else:
                    continue;
            if math.floor((dayValue[0][i+2]%10000)/100) != math.floor((dayValue[0][i+1]%10000)/100):
                for j1 in range(daym-1):
                    avermon[j1+1][(i0-1)*12+j+1] = avermon[j1+1][(i0-1)*12+j+1]/i1;
                i1 = 0;
            else:
                continue;
        if dayValue[0][i+1]%10000 == 1231:
            i0 = i0 +1;
        else:
            continue;
    return avermon

def percentage(avermon):
    [avermonm, avermonn] = avermon.shape;
    perc = np.zeros((avermonm,avermonn),dtype=(float));
    for i in range(avermonn-1):
        sump = 0;
        if avermon[0][i+1] != 0:
            for j in range(avermonm-1):
                sump = sump + avermon[j+1][i+1];
                for j in range(avermonm-1):
                    perc[j+1][i+1] = avermon[j+1][i+1]/sump*100;
                perc[0,...] = avermon[0,...];
                perc[...,0] = avermon[...,0];
        else:
            continue;
    return perc

#写入文件函数
def save(data, path):
    f = xlwt.Workbook(); # 创建工作簿
    sheet1 = f.add_sheet(u'sheet1', cell_overwrite_ok=True) # 创建sheet
    [h, l] = data.shape # h为行数，l为列数
    for i in range(h):
        for j in range(l):
            sheet1.write(i, j, data[i, j]);
    f.save(path);




#读取所需处理站点文件，返回需要读取的站点
def siteread(pathsite):   # path为想要读取的excel文档
    file = openpyxl.load_workbook(pathsite); #打开想要读取的excel文档
    sheet1 = file['Sheet1'];
    m = sheet1.max_row;
    n = sheet1.max_column;
    pathr = [];
    for i in range(n):
        if sheet1.cell(1,i+1).value == 'fluxnetid':
            for j in range(m-1):
                pathr.append(sheet1.cell(j+2,i+1).value);
        elif sheet1.cell(1,i+1).value == 'koeppen_climate':
            a = sheet1.cell(2,i+1).value;
        else:
            continue;
    return pathr, a

#第二版程序，不需要将输入文件分在不同的文件夹里，该程序可以根据输入excel信息进行自动的读取
#需要修改的数据仅有不同的path
path = 'D:/Data/fluxnet/OriginalData/AllHourlyData/';  #储存原始数据的位置
pathinfo = 'D:/Data/fluxnet/TreatedData/ClassLandcover/LandCoverinfo/'  #储存分类信息excel的文件夹
for landinfo in os.listdir(pathinfo):
    pathsite = 'D:/Data/fluxnet/TreatedData/ClassLandcover/LandCoverinfo/' + landinfo;  
    pathr,a = siteread(pathsite);
    #pathsite = 'F:/FLUXnet/TreatedFluxNet/FluxnetInformation/fluxnet_site_info_all.xlsx'  #想要读取的站点信息excel文档，需经过筛选
    os.mkdir(r'D:/Data/fluxnet/TreatedData/SpacificClimateClass/Respeiration/MonthData/' + a);  #创建月平均数据的文件夹
    os.mkdir(r'D:/Data/fluxnet/TreatedData/SpacificClimateClass/Respeiration/MonthPercentage/' + a);    #创建月平均数据百分比的文件夹
    path0 = 'D:/Data/fluxnet/TreatedData/SpacificClimateClass/Respeiration/MonthData/' + a +'/';  #储存计算得到的原始数据月平均值文件的位置
    path1 = 'D:/Data/fluxnet/TreatedData/SpacificClimateClass/Respeiration/MonthPercentage/'+ a +'/';  #储存百分比数据的位置
    os.mkdir(r'D:/Data/fluxnet/TreatedData/SpacificClimateClass/Temperature/MonthData/' + a);
    os.mkdir(r'D:/Data/fluxnet/TreatedData/SpacificClimateClass/Temperature/MonthPercentage/' + a);
    path00 = 'D:/Data/fluxnet/TreatedData/SpacificClimateClass/Temperature/MonthData/' + a +'/';
    path11 = 'D:/Data/fluxnet/TreatedData/SpacificClimateClass/Temperature/MonthPercentage/' + a +'/';
    m = len(pathr);
    for csv_file in os.listdir(path):
        for i in range(m):
            if csv_file[4:10] == pathr[i]:
                
                dayDatae = dayValue(path + csv_file,'RECO_NT_VUT_REF');####
                avermone = avermonth(dayDatae);                        #统计生态系统呼吸的变量
                perce = percentage(avermone);                          ####
                
                dayDatat = dayValue(path + csv_file,'TA_F_MDS');       ####
                avermont = avermonth(dayDatat);                        #统计温度变量
                perct = percentage(avermont);                          ####
                
                path2 = path0 + 'Avermon_' + csv_file[4:10] +'.xls';
                path3 = path1 + 'Perc_' + csv_file[4:10] +'.xls';
                path21= path00 + 'Avermon_' + csv_file[4:10] +'.xls';
                path31 = path11 + 'Perc_' + csv_file[4:10] +'.xls';
                save(avermone,path2);
                save(perce,path3);
                save(avermont,path21);
                save(perct,path31);


'''
#第一版程序，需要将原始文件手动剔出放到指定问价夹下读取
path = 'C:/Users/Lenovo/Desktop/fluxnetdata/TemperateOriginal/'
path1 = 'C:/Users/Lenovo/Desktop/fluxnetdata/temperatePercentageMonthly/'
for csv_file in os.listdir(path):
    dayData = dayValue(path + csv_file);
    avermon = avermonth(dayData);
    perc = percentage(avermon);
    path2 = path1 + 'Perc_' + csv_file;
    save(perc,path2);
'''

    
        

