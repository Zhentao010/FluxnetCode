# -*- coding: utf-8 -*-
"""
Created on Mon Nov 21 21:36:34 2022

@author: Lenovo
"""

'''
统计同一VPD下的Ttemperature和RESP数据,通过不同的vpd计算得到不同的温度敏感性，看温度敏感性的变化
'''

import pandas as pd
import os
import numpy as np
from sklearn import linear_model 
import openpyxl


def TREdata(fileN):
    eco = pd.read_csv(fileN);
    ecom = len(eco);
    data = np.zeros((1,3),dtype = float);
    a = np.zeros((1,3));
    n1 = 0;
    for i1 in range(ecom):
       if eco['TA_F_MDS'][i1]!=-9999 and eco['VPD_F_MDS'][i1]!=-9999 and eco['RECO_NT_VUT_REF'][i1]!=-9999 and eco['RECO_NT_VUT_REF'][i1]!=0:
           data[n1][0] = eco['VPD_F_MDS'][i1];
           data[n1][1] = 1000/(eco['TA_F_MDS'][i1]+273);
           data[n1][2] = np.log(eco['RECO_NT_VUT_REF'][i1]);
           data = np.row_stack((data,a));
           n1 = n1 + 1;
    return data


def value(data,vpd0,vpdn,dVPD,dvpd):  
    
    l = int((vpdn - vpd0)/dVPD + 1);
    ecom = len(data);
    value = np.zeros((10,2*l),dtype = float);
    a = np.zeros((1,2*l));

    vpd = vpd0;
    for i0 in range(l):
        value[0][i0*2] = vpd;
        vpd = vpd + dVPD;
    
    for i1 in range(ecom):
        for i2 in range(l):
            if data[i1][0] >= value[0][i2*2] - dvpd/2 and data[i1][0] < value[0][i2*2] + dvpd/2:
                if data[i1][1] != -9999 and data[i1][2] !=0 and data[i1][2] != -9999:
                    value[1][i2*2] = value[1][i2*2] + 1;
                    b = list(value[1,...]);
                    mm = max(b);
                    [m,n] = value.shape;
                    if mm+2 > m:
                        value = np.row_stack((value,a));
                    value[int(value[1][i2*2])+ 1][i2*2] = data[i1][1];
                    value[int(value[1][i2*2])+ 1][i2*2 + 1] = data[i1][2];
    return value

def regression(value):
    [m,n] = value.shape;
    slope = np.zeros((1,int(n/2)));
    for i in range(int(n/2)):
        a = [];
        b = [];
        for i1 in range(m-2):
            if value[i1+2][2*i]!=0 and value[i1+2][2*i+1]!=0:
                a.append(value[i1+2][2*i]);
                b.append(value[i1+2][2*i+1]);
        regr = linear_model.LinearRegression();
        a = np.array(a);
        b = np.array(b);
        if len(a) != 0:
            regr.fit(a.reshape(-1,1),b);
            k = regr.coef_;
            slope[0][i] = -k[0]*8.62/100;
    return slope

#----------------------------------------------------------------------------------------------------------------------------------
'''读取所需处理站点文件，返回需要读取的站点'''
def siteread(pathsite):   # path为想要读取的excel文档
    file = openpyxl.load_workbook(pathsite); #打开想要读取的excel文档
    sheet1 = file['Sheet1'];
    m = sheet1.max_row;
    n = sheet1.max_column;
    pathr = [];
    for i in range(n):
        if sheet1.cell(1,i+1).value == 'fluxnetid':
            for j in range(m-1):
                pathr.append(sheet1.cell(j+2,i+1).value);
        elif sheet1.cell(1,i+1).value == 'igbp_land_use':
            a = sheet1.cell(2,i+1).value;
        else:
            continue;
    return pathr, a         #a为植被类型，pathr为站点信息


pathinfo = 'D:/Data/fluxnet/TreatedData/ClassLandcover/LandCoverinfo/';  #储存分类信息excel的文件夹
path = 'D:/Data/fluxnet/OriginalData/AllHourlyData/';  #储存原始数据的位置
path0 = 'C:/Users/Lenovo/Desktop/未划分年份的初级数据/同VPD点精度2/';  #储存计算得到的VPD月平均值文件的位置
path1 = 'C:/Users/Lenovo/Desktop/';


vpd0 = 0;
vpdn = 40;
dVPD = 5;
dvpd = 2;
nvpd = int((vpdn - vpd0)/dVPD + 1); 


for landinfo in os.listdir(pathinfo):
    pathsite = 'D:/Data/fluxnet/TreatedData/ClassLandcover/LandCoverinfo/' + landinfo;
    pathsiteid,a = siteread(pathsite);
    m = len(pathsiteid);
    result = np.zeros((6, nvpd*2));   #第一行温度范围，第二行vpd范围
    addre = np.zeros((1, nvpd*2));

    for i in range(nvpd):
        result[0][i*2] = vpd0 + i*dVPD; 

    for csv_file in os.listdir(path):
        for i in range(m):
            if csv_file[4:10] == pathsiteid[i]:
                dTRE = TREdata(path + csv_file);        ##读取原始数据
                dSameVPD = value(dTRE, vpd0, vpdn, dVPD, dvpd);       ##统计同一VPD下的T和RESP
                for i1 in range(nvpd):
                    [mds, nds] = dSameVPD.shape;
                    for i2 in range(mds-2):
                        if dSameVPD[i2+2][i1*2] != 0 and dSameVPD[i2+2][i1*2+1] != 0 and dSameVPD[i2+2][i1*2+1]>-10:
                            result[1][i1*2] = result[1][i1*2] + 1;    #第三行用来储存该VPD范围有多少数据
                            b = list(result[1,...]);
                            mm = max(b);
                            [mre,nre] = result.shape;
                            if mm + 2 > mre:
                                result = np.row_stack((result,addre));
                            result[ int(result[1][i1*2]) + 2 ][i1*2 ] = 1000/(dSameVPD[i2+2][i1*2] + 273 );
                            result[ int(result[1][i1*2]) + 2 ][i1*2 + 1] = dSameVPD[i2+2][i1*2+1];

    #将统计的数据保存
    path00 = path0 + a + '.csv';
    result_pd = pd.DataFrame(result);
    result_pd.to_csv(path00, index= False, header= False)
    #将统计好的数据计算平均值

    
    
